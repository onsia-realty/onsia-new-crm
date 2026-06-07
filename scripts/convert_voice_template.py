# -*- coding: utf-8 -*-
"""직원이 작성한 '음성번호_수기작성_템플릿.xlsx' → 휴대폰번호 변환
규칙:
  - B3 = 날짜(YY-MM-DD), B4 = 메모(아파트/지역) → 전체 동일 적용
  - A8 이하 = 번호(8자리). 8자리는 010 prepend, 11자리는 그대로
  - 이름 = 미정, 파일 내 중복(정규화) 제거
출력:
  - 콘솔 변환 목록
  - 음성번호_변환결과.xlsx (전화번호 | 이름 | 메모 | 날짜)
사용법: python scripts/convert_voice_template.py [파일명]
"""
import sys, re
import openpyxl
from openpyxl.styles import Font, PatternFill

SRC = sys.argv[1] if len(sys.argv) > 1 else '음성번호_수기작성_템플릿.xlsx'

def normalize(raw):
    n = re.sub(r'\D', '', str(raw) if raw is not None else '')
    if len(n) == 10 and n.startswith('1'):
        n = '0' + n
    if len(n) == 8 and not re.match(r'^1[568]', n):
        n = '010' + n
    ok = len(n) == 11 and n.startswith('01')
    return n, ok

def fmt(n):
    return f'{n[:3]}-{n[3:7]}-{n[7:]}' if len(n) == 11 else n

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['번호작성'] if '번호작성' in wb.sheetnames else wb.active

date = (ws['B3'].value or '').__str__().strip()
memo = (ws['B4'].value or '').__str__().strip()

valid, invalid, seen = [], [], set()
for r in range(8, ws.max_row + 1):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    if a is None or str(a).strip() == '':
        continue
    if str(a).strip() == '7892-0012' and b and '예시' in str(b):  # 예시행 skip
        continue
    n, ok = normalize(a)
    if not ok:
        invalid.append((r, str(a).strip(), f'{len(re.sub(chr(92)+"D","",str(a)))}자리'))
        continue
    if n in seen:
        continue
    seen.add(n)
    valid.append(n)
wb.close()

print(f'날짜: {date or "(미입력)"}   메모: {memo or "(미입력)"}')
print(f'유효 번호: {len(valid)}건   /   무효: {len(invalid)}건')
print('-' * 40)
for i, n in enumerate(valid, 1):
    print(f'{i:3}. {fmt(n)}')
if invalid:
    print('\n[무효/확인필요]')
    for r, raw, why in invalid:
        print(f'  행{r}: "{raw}" ({why})')

# 변환결과 엑셀 (대량등록 호환: 전화번호 | 이름 | 메모)
out = openpyxl.Workbook(); o = out.active; o.title = '변환결과'
o.append(['전화번호', '이름', '메모', '날짜'])
for c in range(1, 5):
    o.cell(1, c).font = Font(bold=True, color='FFFFFF')
    o.cell(1, c).fill = PatternFill('solid', fgColor='305496')
for n in valid:
    o.append([fmt(n), '미정', memo, date])
o.column_dimensions['A'].width = 16
o.column_dimensions['B'].width = 8
o.column_dimensions['C'].width = 22
o.column_dimensions['D'].width = 12
o.freeze_panes = 'A2'
out.save('음성번호_변환결과.xlsx')
print('\nsaved 음성번호_변환결과.xlsx')
