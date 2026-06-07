# -*- coding: utf-8 -*-
"""민간임대 파비안.xlsx 정제 → 내부중복/DB중복 대조 → 대량등록용 엑셀 생성
- 전화 8자리(010 생략)는 010 prepend, 11자리는 그대로
- 매칭 기준: 끝 8자리(k8) — 8자리 데이터가 다수라 공통 분모로 안전
- 등록 네이밍(카테고리): 민간임대 3억대
"""
import openpyxl, re, json
from collections import defaultdict, Counter
from openpyxl.styles import Font, PatternFill

SRC = '민간임대 파비안.xlsx'
SITE = '민간임대 3억대'

def digits(v):
    return re.sub(r'\D', '', str(v)) if v is not None else ''

def k8(d):
    return d[-8:] if len(d) >= 8 else None

def normalize(raw):
    """반환 (정규화번호, 유효여부, 사유)"""
    if raw is None or str(raw).strip() == '':
        return ('', False, '번호없음')
    n = digits(raw)
    # 0100xxxx 오타 보정
    if n.startswith('0100') and len(n) == 12:
        n = '010' + n[4:]
    if len(n) == 10 and n.startswith('1'):
        n = '0' + n
    if len(n) == 8 and not re.match(r'^1[568]', n):
        n = '010' + n
    if len(n) == 11 and n.startswith('010'):
        return (n, True, '')
    if len(n) == 11:  # 011 등 구형
        return (n, True, '')
    return (n, False, f'{len(digits(raw))}자리(비정상)')

def fmt(n):
    if len(n) == 11:
        return f'{n[:3]}-{n[3:7]}-{n[7:]}'
    return n

# ---- 파일 읽기 ----
wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
ws = wb['Sheet1']
rows = []
for r in ws.iter_rows(min_row=1, values_only=True):
    idx, name, phone = (list(r) + [None, None, None])[:3]
    if name is None and phone is None:
        continue
    nm = str(name).strip() if name is not None else ''
    n, ok, why = normalize(phone)
    rows.append({'name': nm, 'phone_raw': str(phone).strip() if phone is not None else '',
                 'norm': n, 'ok': ok, 'why': why, 'k8': k8(n) if ok else None})
wb.close()

valid = [x for x in rows if x['ok']]
invalid = [x for x in rows if not x['ok']]

# ---- 파일 내부 중복 (k8 기준) ----
by = defaultdict(list)
for x in valid:
    by[x['k8']].append(x)
dup_groups = {k: v for k, v in by.items() if len(v) > 1}

# ---- DB 대조 ----
db = json.load(open('scripts/db_customers.json', encoding='utf-8'))
dbk = defaultdict(list)
for c in db:
    k = k8(digits(c.get('phone')))
    if k:
        dbk[k].append({'name': c.get('name') or '', 'del': c.get('isDeleted'),
                       'site': c.get('assignedSite') or ''})

# ---- 신규/중복 분류 ----
seen = set()
new_unique = []
dup_in_db = []
dup_in_file = []
for x in valid:
    if x['k8'] in dbk:
        dup_in_db.append(x); continue
    if x['k8'] in seen:
        dup_in_file.append(x); continue
    seen.add(x['k8']); new_unique.append(x)

# ---- 요약 ----
out = []
out.append('=== 민간임대 파비안.xlsx 분석 ===')
out.append(f'총 데이터 행: {len(rows)}')
out.append(f'번호 유효: {len(valid)}  /  비정상·없음(등록불가): {len(invalid)}')
out.append(f'  - 비정상 사유: ' + ', '.join(f'{k}:{v}' for k, v in
           Counter(x['why'] for x in invalid).most_common()))
out.append(f'유효 고유번호(k8): {len(by)}  /  파일내 중복그룹: {len(dup_groups)} (중복행 {sum(len(v) for v in dup_groups.values())})')
out.append('')
out.append(f'DB 기존존재(끝8 매칭): {len(dup_in_db)}')
db_active = sum(1 for x in dup_in_db if any(not m['del'] for m in dbk[x['k8']]))
out.append(f'  - 그중 DB 활성고객: {db_active}  /  삭제(soft-deleted)만 존재: {len(dup_in_db) - db_active}')
out.append('')
out.append(f'★ 신규 등록대상(유니크, DB없음): {len(new_unique)} ★')
out.append(f'(제외) DB중복: {len(dup_in_db)}  +  파일내중복 추가행: {len(dup_in_file)}  +  비정상번호: {len(invalid)}')
out.append('')
# 매칭된 DB 사이트 분포
sc = Counter()
for x in dup_in_db:
    for m in dbk[x['k8']]:
        sc[m['site'] or '(미지정)'] += 1
out.append('DB중복 매칭 사이트 분포 top: ' + ', '.join(f'{s}:{c}' for s, c in sc.most_common(10)))

open('_fabian_summary.txt', 'w', encoding='utf-8').write('\n'.join(out))

# ---- 대량등록용 엑셀 (전화번호 | 이름 | 메모) ----
def write_import(fname, data):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '고객'
    ws.append(['전화번호', '이름', '메모'])
    for c in range(1, 4):
        ws.cell(1, c).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, c).fill = PatternFill('solid', fgColor='305496')
    for x in data:
        ws.append([fmt(x['norm']), x['name'], SITE])
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.freeze_panes = 'A2'
    wb.save(fname)

write_import('민간임대3억대_대량등록_신규.xlsx', new_unique)

# ---- 등록용 JSON ----
to_reg = [{'phone': x['norm'], 'name': x['name'], 'site': SITE} for x in new_unique]
json.dump(to_reg, open('scripts/fabian_to_register.json', 'w', encoding='utf-8'), ensure_ascii=False)

# ---- 검토용 엑셀: 제외/중복 ----
wb = openpyxl.Workbook()
s1 = wb.active; s1.title = 'DB기존중복'
s1.append(['이름', '전화번호', 'DB매칭고객명', 'DB사이트', '활성/삭제'])
for x in dup_in_db:
    m = dbk[x['k8']]
    active = '활성' if any(not mm['del'] for mm in m) else '삭제됨'
    s1.append([x['name'], fmt(x['norm']),
               ', '.join(sorted({mm['name'] for mm in m if mm['name']}))[:40],
               ', '.join(sorted({mm['site'] for mm in m if mm['site']}))[:40], active])
s2 = wb.create_sheet('파일내중복')
s2.append(['전화번호', '건수', '이름들'])
for k, v in sorted(dup_groups.items(), key=lambda kv: -len(kv[1])):
    s2.append([fmt(v[0]['norm']), len(v), ', '.join(x['name'] or '-' for x in v)])
s3 = wb.create_sheet('비정상번호')
s3.append(['이름', '원본번호', '정규화', '사유'])
for x in invalid:
    s3.append([x['name'], x['phone_raw'], x['norm'], x['why']])
for ws in wb.worksheets:
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, c).fill = PatternFill('solid', fgColor='C00000')
    ws.freeze_panes = 'A2'
wb.save('민간임대3억대_제외_검토내역.xlsx')

print('\n'.join(out))
