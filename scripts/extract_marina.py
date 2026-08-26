# -*- coding: utf-8 -*-
"""마리나큐브 원본 엑셀 → scripts/marina_cube.txt (탭 구분: 이름<TAB>전화)

사용: python scripts/extract_marina.py [--dump]
  --dump : 시트 구조/샘플만 출력하고 종료
"""
import io
import sys
import openpyxl

SRC_MAIN = 'D:/DB/성진디비/야목/마리나큐브.xlsx'
SRC_JEONGRI = 'D:/DB/성진디비/야목/마리나큐브 정리.xlsx'
OUT = 'scripts/marina_cube.txt'

log = io.open('scripts/marina_extract.log', 'w', encoding='utf-8')


def w(msg):
    log.write(msg + '\n')
    log.flush()


def dump(path, max_rows=8):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    w('== FILE %s' % path)
    for ws in wb.worksheets:
        w('  sheet=%s rows=%s cols=%s' % (ws.title, ws.max_row, ws.max_column))
        n = 0
        for row in ws.iter_rows(min_row=1, values_only=True):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            w('    %r' % (row,))
            n += 1
            if n >= max_rows:
                break
    wb.close()


def rows_of(path, sheet_names=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        if sheet_names and ws.title not in sheet_names:
            continue
        for row in ws.iter_rows(values_only=True):
            if all(c is None or str(c).strip() == '' for c in row):
                continue
            yield ws.title, row
    wb.close()


def digits(v):
    if v is None:
        return ''
    return ''.join(ch for ch in str(v) if ch.isdigit())


def clean_name(v):
    if v is None:
        return ''
    s = str(v).strip()
    if s in ('-', '_', '.', '*'):
        return ''
    return s


def extract():
    records = []          # (name, phone, source, flags)
    flagged = []          # 삭제요청/결번 표시된 행

    # 1) 마리나큐브.xlsx  Sheet1 (Sheet2는 동일 사본 → 하나만 사용)
    n_main = 0
    for sheet, row in rows_of(SRC_MAIN, sheet_names={'Sheet1'}):
        name, phone = clean_name(row[0]), digits(row[1] if len(row) > 1 else None)
        if not phone or phone == digits('연락처'):
            continue
        if name == '성명':
            continue
        n_main += 1
        records.append((name, phone, 'main'))
    w('마리나큐브.xlsx Sheet1 데이터행: %d' % n_main)

    # 2) 마리나큐브 정리.xlsx Sheet1 (0=현장, 1=이름, 2=전화, 7=삭제요청, 8=부재, 9=결번)
    n_j = 0
    for sheet, row in rows_of(SRC_JEONGRI, sheet_names={'Sheet1'}):
        site = clean_name(row[0])
        if site != '마리나큐브':
            continue
        name = clean_name(row[1] if len(row) > 1 else None)
        phone = digits(row[2] if len(row) > 2 else None)
        if not phone:
            continue
        del_req = row[7] if len(row) > 7 else None
        dead = row[9] if len(row) > 9 else None
        n_j += 1
        if (del_req is not None and str(del_req).strip()) or (dead is not None and str(dead).strip()):
            flagged.append((name, phone, str(del_req), str(dead)))
            continue
        records.append((name, phone, 'jeongri'))
    w('마리나큐브 정리.xlsx 마리나큐브 행: %d (삭제요청/결번 제외 %d건)' % (n_j, len(flagged)))
    for f in flagged[:50]:
        w('  제외(플래그): %r' % (f,))

    # 3) 합치기 + 중복 제거 (첫 등장 우선, 이름 있는 쪽 우선)
    best = {}
    order = []
    for name, phone, src in records:
        if phone not in best:
            best[phone] = [name, src]
            order.append(phone)
        else:
            if not best[phone][0] and name:
                best[phone][0] = name

    w('총 수집: %d행 / 고유 번호: %d건 (리스트 내 중복 %d건 제거)'
      % (len(records), len(order), len(records) - len(order)))

    valid, invalid = [], []
    for phone in order:
        if len(phone) == 11 and phone.startswith('010') and phone[3] not in '01':
            valid.append(phone)
        else:
            invalid.append(phone)
    w('형식 이상(제외): %d건' % len(invalid))
    for p in invalid:
        w('  이상번호: %s | %s' % (p, best[p][0]))

    with io.open(OUT, 'w', encoding='utf-8') as f:
        for phone in valid:
            f.write('%s\t%s\n' % (best[phone][0], phone))
    w('저장: %s (%d건)' % (OUT, len(valid)))
    return len(valid)


if __name__ == '__main__':
    if '--dump' in sys.argv:
        dump(SRC_MAIN)
        dump(SRC_JEONGRI)
        print('dumped to scripts/marina_extract.log')
        sys.exit(0)
    n = extract()
    print('OK: %d rows -> %s (log: scripts/marina_extract.log)' % (n, OUT))
