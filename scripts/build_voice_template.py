# -*- coding: utf-8 -*-
"""음성파일 번호 수기 작성용 엑셀 템플릿 생성
구조:
  - 날짜(1회 입력, YY-MM-DD), 메모=아파트/지역(1회 입력 → 전체 동일 적용)
  - 번호(8자리)를 한 줄에 하나씩 수기 작성 (예 7892-0012)
업로드 시 변환규칙: 8자리 → 010-XXXX-XXXX, 이름=미정, 메모/날짜 일괄 적용
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '번호작성'

THIN = Side(style='thin', color='BBBBBB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', fgColor='305496')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')   # 노란 = 직원 입력란
EX_FONT = Font(color='999999', italic=True)

# ── 타이틀 ──
ws.merge_cells('A1:C1')
ws['A1'] = '📞 음성파일 번호 수기 작성 템플릿'
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws['A1'].fill = HEADER_FILL
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

# ── 상단 공통 입력(1회 작성) ──
ws['A3'] = '날짜'
ws['A3'].font = Font(bold=True)
ws['B3'] = ''                      # 예: 26-06-07
ws['B3'].fill = INPUT_FILL
ws['B3'].border = BORDER
ws['C3'] = '← YY-MM-DD 형식 (예: 26-06-07)'
ws['C3'].font = EX_FONT

ws['A4'] = '메모(아파트/지역)'
ws['A4'].font = Font(bold=True)
ws['B4'] = ''                      # 예: 안산 OO아파트
ws['B4'].fill = INPUT_FILL
ws['B4'].border = BORDER
ws['C4'] = '← 한 번만 작성하면 전체 번호에 동일하게 적용됩니다 (예: 안산, 용인, OO아파트)'
ws['C4'].font = EX_FONT

# ── 안내 ──
ws.merge_cells('A6:C6')
ws['A6'] = ('▶ 음성파일을 들으며 아래 "번호(8자리)" 칸에 한 줄에 하나씩 적어주세요. '
            '예시 7892-0012 처럼 8자리만 적으면 됩니다. (010은 자동으로 붙습니다)')
ws['A6'].font = Font(size=10, color='C00000')
ws['A6'].alignment = Alignment(wrap_text=True, vertical='center')
ws.row_dimensions[6].height = 30

# ── 번호 입력 헤더 ──
ws['A7'] = '번호(8자리)'
ws['A7'].font = Font(bold=True, color='FFFFFF')
ws['A7'].fill = HEADER_FILL
ws['A7'].alignment = Alignment(horizontal='center')
ws['A7'].border = BORDER

# 예시 행(회색, 직원이 지우고 작성)
ws['A8'] = '7892-0012'
ws['A8'].font = EX_FONT
ws['A8'].alignment = Alignment(horizontal='center')
ws['A8'].border = BORDER
ws['B8'] = '← 예시 (지우고 작성)'
ws['B8'].font = EX_FONT

# 빈 입력칸 200줄 (노란 배경)
for r in range(9, 209):
    cell = ws.cell(r, 1)
    cell.fill = INPUT_FILL
    cell.border = BORDER
    cell.alignment = Alignment(horizontal='center')

# 너비
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 50
ws.freeze_panes = 'A8'

wb.save('음성번호_수기작성_템플릿.xlsx')
print('saved 음성번호_수기작성_템플릿.xlsx')
